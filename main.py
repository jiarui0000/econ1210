import openpyxl
import numpy


def equation_transform(mode, a, c):
    # mode = int(input("mode 1: P = c + aQ -> Q = c + aP\n mode 2: Q = c + aP -> P = c + aQ\n mode: "))
    c1 = -c / a
    a1 = 1 / a
    if mode == 1:  # mode 1: P = c + aQ -> Q = c + aP
        print("P =", round(c, 4), "+(", round(a, 4), ")Q -> Q =", round(c1, 4), "+(", round(a1, 4), ")P. ")
    elif mode == 2:  # mode 2: Q = c + aP -> P = c + aQ
        print("Q =", round(c, 4), "+(", round(a, 4), ")P -> P =", round(c1, 4), "+(", round(a1, 4), ")Q. ")


def worker_allocate():
    book0 = openpyxl.load_workbook("worker allocate.xlsx")
    sheet0 = book0.worksheets[0]
    r = sheet0.max_row
    c = sheet0.max_column

    queue = []
    newSheet = [[], []]
    newSheet1 = [[], []]
    nw = sheet0.cell(row=1, column=1).value
    if nw == -1:
        nw = (r - 1) * (c - 1)
    for i in range(2, c + 1):
        newSheet.append([])
        newSheet1.append([])
        newSheet[i].append(i)
        newSheet1[i].append(-1)
        for j in range(2, r + 1):
            if j != 2:
                newSheet[i].append(sheet0.cell(row=j, column=i).value - sheet0.cell(row=j - 1, column=i).value)
            else:
                newSheet[i].append(sheet0.cell(row=j, column=i).value)
            newSheet1[i].append(-1)

    ptr = [-1, -1]
    for i in range(2, c + 1):
        ptr.append(1)
    while len(queue) < nw:
        max_temp = 0
        for i in range(1, c + 1):
            if ptr[i] != -1:
                max_temp = newSheet[2][ptr[i]]
                break
        choice = 2
        for i in range(2, c + 1):
            if ptr[i] != -1:
                if newSheet[i][ptr[i]] > max_temp:
                    choice = i
                    max_temp = newSheet[i][ptr[i]]
        ptr[choice] += 1
        queue.append((choice, ptr[choice] - 1))
        newSheet1[choice][ptr[choice] - 1] = len(queue)
        if ptr[choice] >= r:
            ptr[choice] = -1

    if len(book0.worksheets) > 1:
        book0.remove(book0.worksheets[1])
    sheet1 = book0.create_sheet("output")

    for i in range(1, c + 1):
        sheet1.cell(1, i, sheet0.cell(1, i).value)
    for i in range(1, r):
        sheet1.cell(i + 1, 1, str(i))
        for j in range(2, c + 1):
            sheet1.cell(i + 1, j, str(newSheet[j][i]) + "(" + str(newSheet1[j][i]) + ")")

    for i in range(2, c + 1):
        if ptr[i] != -1:
            print(ptr[i] - 1, end="\t")
        else:
            print(r - 1, end="\t")
    print()

    book0.save("worker allocate.xlsx")


def TB_calculate(i, t=(0, 0, 0, 0)):
    a, b, c, d = t
    f = a * numpy.square(i) + b * i + c + d * numpy.sqrt(i)
    return f


def MB_calculate():
    print("mode 1: total; mode 2: average; mode 3: 直接比较MC;")
    mode = int(input("mode="))
    a = int(input("二次"))
    b = int(input("一次"))
    c = int(input("常数"))
    d = int(input("sqrt"))
    t = (a, b, c, d)
    if mode == 1:  # calculate through total benefit
        m = int(input("猜测最大"))
        print("from total benefit")
        for i in range(m + 1):
            print(i, ":", TB_calculate(i, t) - TB_calculate(i - 1, t))
    elif mode == 2:  # calculate through average benefit
        m = int(input("猜测最大"))
        print("from average benefit")
        for i in range(m + 1):
            print(i, ":", i * TB_calculate(i, t) - (i - 1) * TB_calculate(i - 1, t))
    elif mode == 3:
        mc = int(input("MC"))
        i = 1
        while (TB_calculate(i, t) - TB_calculate(i - 1, t)) >= mc:
            i += 1
        i -= 1
        print("TIME=", i, ", MB=", TB_calculate(i, t) - TB_calculate(i - 1, t))


def comparative_advantage():
    mode = int(input("mode 1: 单位花费产量; mode 2: 单位产量花费; \\t分割数据输入\nmode: "))
    n = 2
    c = []
    for i in range(n):
        c.append(input().split("\t"))

    if mode == 1:  # 条件为单位花费产量
        for i in range(n):
            print(round(float(c[i][1]) / float(c[i][0]), 4), "\t", round(float(c[i][0]) / float(c[i][1]), 4))
    elif mode == 2:  # 条件为单位产量花费
        for i in range(n):
            print(round(float(c[i][0]) / float(c[i][1]), 4), "\t", round(float(c[i][1]) / float(c[i][0]), 4))
    print("数字越小越擅长!!! 数字代表 X 另外工作= 1 本工作")


def corporate_production():
    print("单位花费产量; \\t分割数据输入")
    n = int(input("人数: "))
    m = 2
    c = []
    for i in range(n):
        c.append(input().split("\t"))
    d = []
    for j in range(m):
        d.append([])

    for i in range(n):
        d[0].append(round(float(c[i][1]) / float(c[i][0]), 4))
        d[1].append(round(float(c[i][0]) / float(c[i][1]), 4))

    queue = []
    for i in range(m):
        queue.append([])
        for j in range(n):
            queue[i].append(d[i].index(min(d[i])))
            d[i][d[i].index(min(d[i]))] = 1000

    print(queue[0])
    for i in range(n + 1):
        x1 = 0
        x2 = 0
        for j in range(i):
            x1 += float(c[queue[0][j]][0])
        for j in range(i, n):
            x2 += float(c[queue[0][j]][1])
        print("(", round(x1, 4), ",", round(x2, 4), ")")


def judge_elasticity(e):
    if numpy.abs(e - 1) < 0.00000001:
        print("unitary elastic")
    elif numpy.abs(e) < 1:
        print("inelastic")
    else:
        print("elastic")


def midpoint_elasticity():
    p1 = float(input("P1="))
    q1 = float(input("Q1="))
    p2 = float(input("P2="))
    q2 = float(input("Q2="))
    midp = (p1 + p2) / 2
    midq = (q1 + q2) / 2
    if p1 == p2:
        print("无穷. elastic")
        return
    if q1 == q2:
        print("0. inelastic")
        return
    e = ((q1 - q2) / midq) / ((p1 - p2) / midp)
    print("elasticity =", round(e, 4))
    judge_elasticity(e)


def curve_elasticity():
    mode = input("calculate elasticity through curve equation. \nmode 1: P = c + aQ\nmode 2: Q = c + aP\nmode=")
    p = 1
    q = 1
    e = -1
    if mode == "1":  # P = c + aQ
        print("P=c+aQ")
        c = float(input("c="))
        a = float(input("a="))
        s = input("Given P or Q? p/q ")
        if s == "p":
            p = float(input("P="))
            q = (p - c) / a
        elif s == "q":
            q = float(input("Q="))
            p = c + a * q
        e = 1 / a * p / q
    elif mode == "2":  # Q = c + aP
        print("Q=c+aP")
        c = float(input("c="))
        a = float(input("a="))
        s = input("Given P or Q? p/q")
        if s == "p":
            p = float(input("P="))
            q = c + a * p
        elif s == "q":
            q = float(input("Q="))
            p = (q - c) / a
        e = a * p / q
    print("elasticity = ", round(e, 4))
    judge_elasticity(e)


def price_change():
    print("calculate influence of percentage change")
    print("in demand / supply? d/s")
    mode = input("mode: ")
    s = float(input("price elasticity of supply: "))
    d = float(input("price elasticity of demand: "))
    e0 = numpy.abs(s) + numpy.abs(d)
    dp, dq = (0, 0)
    if mode == "d":
        dd = float(input("increase in demand:(百分比,负数为减少) "))
        dp = dd / e0
        dq = dp * s
    elif mode == "s":
        ds = float(input("increase in supply:(百分比,负数为减少) "))
        dp = -ds / e0
        dq = dp * d
    print("change in price = %", round(dp, 4))
    print("change in quantity traded = %", round(dq, 4))


def tax_influence():
    print("calculate change caused by tax")
    s = input("P=/Q=? p/q ")
    pd, ps, q, p0, q0, tr, dwl = (0, 0, 0, 0, 0, 0, 0)
    if s == "p":
        print("Pd = c1 + a1*Q")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Ps = c2 + a2*Q")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        t = float(input("tax? (subsidy -): "))
        q = (t - c1 + c2) / (a1 - a2)
        pd = c1 + a1 * q
        ps = c2 + a2 * q
        tr = t * q
        q0 = (c2 - c1) / (a1 - a2)
        p0 = c1 + a1 * q0
        dwl = t * (q0 - q) / 2
    elif s == "q":
        print("Qd = c1 + a1*P")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Qs = c2 + a2*P")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        t = float(input("tax? (subsidy -): "))
        ps = (c2 - c1 - a1 * t) / (a1 - a2)
        pd = ps + t
        q = c2 + a2 * ps
        tr = t * q
        p0 = (c1 - c2) / (a2 - a1)
        q0 = c1 + a1 * p0
        dwl = t * (q0 - q) / 2
    print("Ordinary (P, Q)= (", round(p0, 4), ",", round(q0, 4), ")")
    print("new quantity=", round(q, 4), "; q-q0=", round(q - q0, 4))
    print("demand price=", round(pd, 4), "; pd-p0=", round(pd - p0, 4))
    print("supply price=", round(ps, 4), "; ps-p0=", round(ps - p0, 4))
    print("tax revenue=", round(tr, 4))
    print("welfare loss=", round(dwl, 4))


def loss_to_tax():
    print("calculate tax/subsidy by welfare loss")
    dwl = float(input("welfare loss="))
    c = input("tax/subsidy? t/s ")
    s = input("P=/Q=? p/q ")
    pd, ps, q, p0, q0, tr, dq = (0, 0, 0, 0, 0, 0, 0)
    if (s == "p") & (c == "t"):
        print("Pd = c1 + a1*Q")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Ps = c2 + a2*Q")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        q0 = (c2 - c1) / (a1 - a2)
        p0 = c1 + a1 * q0
        dq = numpy.sqrt(2*dwl/(a2-a1))
        q = q0 - dq
        t = dq*(a2 - a1)
        tr = t * q
        pd = c1 + a1 * q
        ps = c2 + a2 * q
    elif (s == "q") & (c == "t"):
        print("Qd = c1 + a1*P")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Qs = c2 + a2*P")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        p0 = (c1 - c2) / (a2 - a1)
        q0 = c1 + a1 * p0
        dq = numpy.sqrt(2*dwl*(a1*a2)/(a1-a2))
        q = q0 - dq
        t = dq * (a1-a2)/(a1*a2)
        ps = (c2 - c1 - a1 * t) / (a1 - a2)
        pd = ps + t
        tr = t * q
    elif (s == "p") & (c == "s"):
        print("Pd = c1 + a1*Q")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Ps = c2 + a2*Q")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        q0 = (c2 - c1) / (a1 - a2)
        p0 = c1 + a1 * q0
        dq = numpy.sqrt(2*dwl/(a2-a1))
        q = q0 + dq
        t = -dq*(a2 - a1)
        tr = t * q
        pd = c1 + a1 * q
        ps = c2 + a2 * q
    elif (s == "q") & (c == "s"):
        print("Qd = c1 + a1*P")
        c1 = float(input("c1="))
        a1 = float(input("a1="))
        print("Qs = c2 + a2*P")
        c2 = float(input("c2="))
        a2 = float(input("a2="))
        p0 = (c1 - c2) / (a2 - a1)
        q0 = c1 + a1 * p0
        dq = numpy.sqrt(2*dwl*(a1*a2)/(a1-a2))
        q = q0 + dq
        t = -dq * (a1-a2)/(a1*a2)
        ps = (c2 - c1 - a1 * t) / (a1 - a2)
        pd = ps + t
        tr = t * q
    print("Ordinary (P, Q)= (", round(p0, 4), ",", round(q0, 4), ")")
    print("new quantity=", round(q, 4), "; q-q0=", round(q - q0, 4))
    print("demand price=", round(pd, 4), "; pd-p0=", round(pd - p0, 4))
    print("supply price=", round(ps, 4), "; ps-p0=", round(ps - p0, 4))
    print("tax revenue=", round(tr, 4))
    print("tax (subsidy -) :", round(t, 4))


# equation_transform(mode=1, c=-20000, a=75)  # 等式变形; (mode, c, a); mode 1: P=; mode 2: Q=;
# worker_allocate()  # 人员分配达到最大产量; io: "worker allocate.xlsx"; 1行1列是总人数; 输入-1全排;
# MB_calculate()  # 计算marginal benefit;
# TB_calculate(5, (1, 2, 3, 4))  # 计算total benefit; (次数, (2次,1次,常数,根号));
# comparative_advantage()  # 相对工作效率换算;
# corporate_production()  # 计算合作曲线上的点坐标;
# judge_elasticity(e=-1.5)  # 根据elasticity的值判断是否为elastic
# midpoint_elasticity()  # 由mid-point法计算elasticity
# curve_elasticity()  # 根据表达式计算elasticity
# price_change()  # 计算demand/supply变化带来的影响
# tax_influence()  # 计算收税的结果
loss_to_tax()  # 根据财富损失算收税/补贴
